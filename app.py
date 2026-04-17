import yt_dlp
import pandas as pd
import streamlit as st
import io
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl.utils import get_column_letter
from datetime import datetime

# --- CORE HELPERS ---

def hms(s):
    """Converts seconds to HH:MM:SS."""
    if not s or s < 0: return "00:00:00"
    h = s // 3600
    m = (s % 3600) // 60
    return f"{int(h):02}:{int(m):02}:{int(s % 60):02}"

def clean_name(n, maxlen=150):
    """Sanitizes strings for filenames and sheet names."""
    forbidden = [':', '\\', '/', '?', '*', '[', ']', '<', '>', '|']
    for char in forbidden: n = n.replace(char, '')
    return "".join(c for c in n if c.isalnum() or c in " _-").strip()[:maxlen] or "data"

def clean_sheet(n, used):
    """Ensures Excel sheet names are unique and within length limits."""
    base = clean_name(n, 25)
    name, counter = base, 1
    while name.lower() in used:
        suffix = f"_{counter}"
        name = base[:31-len(suffix)] + suffix
        counter += 1
    used.add(name.lower())
    return name

def auto_adjust_columns(ws, max_width=60):
    """Auto-fits column widths in OpenPyXL worksheets."""
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                length = len(str(cell.value)) if cell.value else 0
                if length > max_length: max_length = length
            except: pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, max_width)

def get_ydl_opts(flat=True):
    """Standard yt-dlp configuration."""
    return {'quiet': True, 'extract_flat': flat, 'skip_download': True, 'ignoreerrors': True, 'ignore_no_formats_error': True}

def normalize_channel_url(url: str) -> str:
    """Cleans YouTube URLs to prevent redundant processing."""
    if not url: return url
    url = url.strip()
    if "?" in url: url = url.split("?")[0]
    return url.rstrip("/")

# --- CACHED DATA FETCHING ---

@st.cache_data(ttl=3600, show_spinner=False)
def get_channel_info_cached(url):
    """Fetches high-level channel metadata."""
    with yt_dlp.YoutubeDL(get_ydl_opts(False)) as d:
        return d.extract_info(url, download=False)

@st.cache_data(ttl=3600, show_spinner=False)
def extract_tab_cached(url, tab):
    """Fetches metadata for Videos, Shorts, or Streams tabs."""
    target_url = url.rstrip('/') + f"/{tab}"
    videos = {}
    with yt_dlp.YoutubeDL(get_ydl_opts(True)) as d:
        try:
            info = d.extract_info(target_url, download=False)
            if info and 'entries' in info:
                for v in info['entries']:
                    if not v: continue
                    vid_id = v.get('id')
                    if not vid_id: continue
                    dur = v.get('duration') or 0
                    date_val = v.get('upload_date')
                    try:
                        dt = datetime.strptime(date_val, "%Y%m%d").date() if date_val else None
                    except: dt = None
                    videos[vid_id] = {
                        "Video Title": v.get('title') or "video",
                        "Duration (HH:MM:SS)": hms(dur),
                        "Video Link": f"https://www.youtube.com/watch?v={vid_id}",
                        "Publish Date": dt,
                        "seconds": dur,
                        "type": tab
                    }
        except: pass
    return videos

@st.cache_data(ttl=3600, show_spinner=False)
def extract_playlist_cached(pl_url, pl_name):
    """Fetches metadata for all videos within a specific playlist."""
    data, total, ids = [], 0, set()
    with yt_dlp.YoutubeDL(get_ydl_opts(True)) as d:
        try:
            info = d.extract_info(pl_url, download=False)
            if info and 'entries' in info:
                for v in info['entries']:
                    if not v: continue
                    vid_id = v.get('id')
                    dur = v.get('duration') or 0
                    date_val = v.get('upload_date')
                    try:
                        dt = datetime.strptime(date_val, "%Y%m%d").date() if date_val else None
                    except: dt = None
                    total += dur
                    if vid_id: ids.add(vid_id)
                    data.append({
                        "Video Title": v.get('title') or "video",
                        "Duration (HH:MM:SS)": hms(dur),
                        "Video Link": f"https://www.youtube.com/watch?v={vid_id}",
                        "Publish Date": dt
                    })
        except: pass
    return pl_name, pl_url, data, total, ids

# --- EXCEL ENGINE ---

def build_excel(results, summary, total_channel, metadata):
    """Generates the formatted Excel file with hyperlinks and styling."""
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl", date_format="YYYY-MM-DD", datetime_format="YYYY-MM-DD") as writer:
        # Metadata Sheet
        pd.DataFrame(list(metadata.items()), columns=["Field", "Value"]).to_excel(writer, sheet_name="Channel_Info", index=False)
        
        # Summary Sheet
        df_sum = pd.DataFrame({
            "Category": [s[1] for s in summary],
            "Video Count": [s[4] for s in summary],
            "Duration": [hms(s[3]) for s in summary],
            "Link": [s[2] for s in summary]
        })
        df_sum.loc[len(df_sum)] = {"Category": "TOTAL CHANNEL DURATION", "Video Count": "", "Duration": hms(total_channel), "Link": ""}
        df_sum.to_excel(writer, sheet_name="Summary", index=False)
        
        wb = writer.book
        # Detailed Sheets
        for sn, n, pl, d, t in results:
            df = pd.DataFrame(d)
            if not df.empty and "Publish Date" in df.columns:
                df = df.sort_values("Publish Date", ascending=False, na_position='last')
            df.to_excel(writer, sheet_name=sn, index=False, startrow=1)
            ws = wb[sn]
            ws["A1"] = "⬅ Back to Summary"
            ws["A1"].hyperlink, ws["A1"].style = "#'Summary'!A1", "Hyperlink"
            for r in range(3, len(d) + 3):
                cell = ws[f"C{r}"]
                if cell.value and str(cell.value).startswith("http"):
                    cell.hyperlink, cell.style = cell.value, "Hyperlink"
            auto_adjust_columns(ws)
            
        auto_adjust_columns(wb["Channel_Info"])
        ws_sum = wb["Summary"]
        for i, r in enumerate(summary, 2):
            ws_sum[f"A{i}"].hyperlink, ws_sum[f"A{i}"].style = f"#'{r[0]}'!A1", "Hyperlink"
            if r[2].startswith("http"):
                ws_sum[f"D{i}"].hyperlink, ws_sum[f"D{i}"].style = r[2], "Hyperlink"
        auto_adjust_columns(ws_sum)
        
    return output.getvalue()

# --- STREAMLIT UI ---

st.set_page_config(page_title="YTCore Archive", layout="wide", page_icon="🎥")
st.title("🎥 YTCore Archive: Full Channel Exporter")

url_input = st.text_input("YouTube Channel URL (e.g., https://www.youtube.com/@Name)")

col_run, col_clear = st.columns([1, 4])
with col_run:
    run_btn = st.button("Generate Archive", type="primary")
with col_clear:
    if st.button("Clear Cache"):
        st.cache_data.clear()
        st.success("Cache cleared!")

if run_btn:
    if not url_input:
        st.error("Please enter a URL")
    else:
        clean_url = normalize_channel_url(url_input)
        progress = st.progress(0.05)
        
        with st.spinner("Processing full channel archive..."):
            try:
                # 1. Channel Info & Setup
                channel_info = get_channel_info_cached(clean_url)
                channel_name = clean_name(channel_info.get("title", "channel"))
                channel_url_full = channel_info.get("webpage_url", clean_url)

                # 2. Get Playlist List
                y_playlists = []
                with yt_dlp.YoutubeDL(get_ydl_opts(True)) as d:
                    info = d.extract_info(clean_url + '/playlists', download=False)
                    if info and 'entries' in info:
                        for e in info['entries']:
                            if e: y_playlists.append({'name': e.get('title') or "playlist", 'url': e.get('url') or e.get('webpage_url')})

                # 3. Get Tabs (Parallel extraction via caching)
                all_content = {}
                for i, tab in enumerate(["videos", "shorts", "streams"]):
                    all_content.update(extract_tab_cached(clean_url, tab))
                    progress.progress(0.20 + (i + 1) * 0.15)

                results, summary, playlist_video_ids, used_sheets = [], [], set(), {"summary", "channel_info"}

                # 4. Extract Playlists (Parallel)
                with ThreadPoolExecutor(max_workers=5) as ex:
                    futures = [ex.submit(extract_playlist_cached, p['url'], p['name']) for p in y_playlists]
                    for f in as_completed(futures):
                        n, pl, d, t, ids = f.result()
                        if not d: continue
                        playlist_video_ids.update(ids)
                        sn = clean_sheet(n, used_sheets)
                        results.append((sn, n, pl, d, t))
                        summary.append((sn, n, pl, t, len(d)))

                # 5. Process Remainder (Videos NOT in playlists)
                for label, tab_type in {"Videos": "videos", "Shorts": "shorts", "Streams": "streams"}.items():
                    items = [v for k, v in all_content.items() if k not in playlist_video_ids and v['type'] == tab_type]
                    if items:
                        sn = clean_sheet(label, used_sheets)
                        dur = sum(i['seconds'] for i in items)
                        results.append((sn, label, clean_url, items, dur))
                        summary.append((sn, label, clean_url, dur, len(items)))

                total_duration = sum(s[3] for s in summary)
                metadata = {
                    "Channel Name": channel_name,
                    "Channel URL": channel_url_full,
                    "Export Tool": "YTCore Archive",
                    "Export Date": datetime.now().strftime("%Y-%m-%d %H:%M"),
                    "Total Duration": hms(total_duration)
                }

                # 6. Build the Excel file
                excel_file = build_excel(results, summary, total_duration, metadata)

                # 7. UI Preview & Download
                st.subheader("📊 Channel Metadata")
                st.table(pd.DataFrame(list(metadata.items()), columns=["Field", "Value"]))

                st.subheader("📋 Content Summary")
                df_summary_preview = pd.DataFrame({
                    "Category": [s[1] for s in summary],
                    "Count": [s[4] for s in summary],
                    "Duration": [hms(s[3]) for s in summary],
                    "Link": [s[2] for s in summary]
                })
                df_summary_preview.loc[len(df_summary_preview)] = {"Category": "TOTAL", "Count": sum(s[4] for s in summary), "Duration": hms(total_duration), "Link": ""}
                st.dataframe(df_summary_preview, use_container_width=True)

                progress.progress(1.0)
                st.success(f"Successfully archived {channel_name}!")
                
                filename = f"{channel_name}_{datetime.now().strftime('%Y%m%d')}.xlsx"
                st.download_button("📥 Download Excel Archive", excel_file, filename, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

            except Exception as e:
                st.error(f"Archive Error: {e}")
